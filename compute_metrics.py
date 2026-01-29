#!/usr/bin/env python3
"""
Aggregate per-method column averages across MULTIPLE completed judging spreadsheets
generated from the same template (create_retrieval_judging_spreadsheet).

For each spreadsheet and each method block:
- <method>_good: TRUE/FALSE -> 1/0; averaged over non-blank cells
- <method>_quality_order: numeric; averaged over non-blank cells

Then aggregates across all spreadsheets by pooling all valid cell values
(not by averaging per-file averages).

Usage:
  python summarize_judging_multi.py comparison_output/a.xlsx comparison_output/b.xlsx
  python summarize_judging_multi.py *.xlsx --sheet Judging
  python summarize_judging_multi.py a.xlsx b.xlsx --out summary.json
  python summarize_judging_multi.py a.xlsx b.xlsx --strict-methods
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class ColInfo:
    method: str
    video_path_col: int
    good_col: int
    quality_order_col: int


def _as_bool01(v: Any) -> Optional[float]:
    """Convert to 0/1 float, or None if blank/invalid."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)) and v in (0, 1):
        return float(int(v))
    if isinstance(v, str):
        s = v.strip().lower()
        if s == "true":
            return 1.0
        if s == "false":
            return 0.0
    return None


def _as_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _mean(xs: List[float]) -> Optional[float]:
    return (sum(xs) / len(xs)) if xs else None


def _parse_template_headers(ws) -> Dict[str, ColInfo]:
    """
    Infer method blocks from headers in row 1:
      rank,
      <m>_video_path, <m>_good, <m>_quality_order,
      <m2>_video_path, ...
    Returns: method -> ColInfo
    """
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = col

    if "rank" not in headers:
        raise ValueError("Missing 'rank' header in row 1.")

    methods = sorted({h[: -len("_video_path")] for h in headers if h.endswith("_video_path")})
    if not methods:
        raise ValueError("No method blocks found (expected headers like '<method>_video_path').")

    out: Dict[str, ColInfo] = {}
    for m in methods:
        vp, gd, qo = f"{m}_video_path", f"{m}_good", f"{m}_quality_order"
        missing = [x for x in (vp, gd, qo) if x not in headers]
        if missing:
            raise ValueError(f"Method '{m}' missing headers: {missing}")
        out[m] = ColInfo(
            method=m,
            video_path_col=headers[vp],
            good_col=headers[gd],
            quality_order_col=headers[qo],
        )
    return out


def iter_values_from_sheet(
    xlsx_path: str,
    sheet_name: str,
    require_video_path: bool,
) -> Dict[str, Dict[str, List[float]]]:
    """
    Returns pooled values from a single sheet:
      {
        method: {
          "good": [0/1...],
          "quality_order": [numbers...]
        }
      }
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{xlsx_path}: sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    cols_by_method = _parse_template_headers(ws)

    out: Dict[str, Dict[str, List[float]]] = {
        m: {"good": [], "quality_order": []} for m in cols_by_method
    }

    for r in range(2, ws.max_row + 1):
        for m, ci in cols_by_method.items():
            vp = ws.cell(row=r, column=ci.video_path_col).value
            if require_video_path and (vp is None or str(vp).strip() == ""):
                continue

            g = _as_bool01(ws.cell(row=r, column=ci.good_col).value)
            if g is not None:
                out[m]["good"].append(g)

            q = _as_float(ws.cell(row=r, column=ci.quality_order_col).value)
            if q is not None:
                out[m]["quality_order"].append(q)

    return out


def summarize_many(
    xlsx_paths: List[str],
    sheet_name: str = "Judging",
    require_video_path: bool = True,
    strict_methods: bool = False,
) -> Dict[str, Dict[str, Any]]:
    """
    Pool all cell values across all spreadsheets, then compute means.

    If strict_methods=False (default):
      - We take the union of methods found across files.
      - If a file is missing a method, it just contributes nothing for that method.

    If strict_methods=True:
      - All files must have the same set of methods (else error).
    """
    pooled_good: Dict[str, List[float]] = {}
    pooled_order: Dict[str, List[float]] = {}

    reference_methods: Optional[set[str]] = None

    for path in xlsx_paths:
        per_file = iter_values_from_sheet(path, sheet_name, require_video_path)

        file_methods = set(per_file.keys())
        if reference_methods is None:
            reference_methods = file_methods
        elif strict_methods and file_methods != reference_methods:
            raise ValueError(
                f"Method mismatch in {path}. Expected {sorted(reference_methods)}, got {sorted(file_methods)}"
            )

        for m, vals in per_file.items():
            pooled_good.setdefault(m, []).extend(vals["good"])
            pooled_order.setdefault(m, []).extend(vals["quality_order"])

    methods = sorted(set(pooled_good.keys()) | set(pooled_order.keys()))
    out: Dict[str, Dict[str, Any]] = {}
    for m in methods:
        gvals = pooled_good.get(m, [])
        ovals = pooled_order.get(m, [])
        out[m] = {
            "good_avg": _mean(gvals),
            "good_count": len(gvals),
            "quality_order_avg": _mean(ovals),
            "quality_order_count": len(ovals),
        }
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="+", help="Paths to completed judging spreadsheets (.xlsx)")
    ap.add_argument("--sheet", default="Judging", help="Sheet name (default: Judging)")
    ap.add_argument(
        "--include-blank-video-path",
        action="store_true",
        help="Also consider rows where <method>_video_path is blank (default: skip).",
    )
    ap.add_argument(
        "--strict-methods",
        action="store_true",
        help="Error if files don't all contain the same method columns.",
    )
    ap.add_argument("--out", default="", help="Optional JSON output path.")
    args = ap.parse_args()

    summary = summarize_many(
        xlsx_paths=args.xlsx,
        sheet_name=args.sheet,
        require_video_path=not args.include_blank_video_path,
        strict_methods=args.strict_methods,
    )

    print(json.dumps(summary, indent=2, sort_keys=True))

    if args.out:
        with open(args.out, "w") as f:
            json.dump(summary, f, indent=2, sort_keys=True)
        print(f"\nWrote: {args.out}")


if __name__ == "__main__":
    main()
