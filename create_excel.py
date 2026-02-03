import json
import os
from typing import List, Dict, Any, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PATH_PREFIX = ""


def _load_json_list(path: str) -> List[Dict[str, Any]]:
    with open(path, "r") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError(f"{path} is not a JSON list.")
    return data


def _normalize_rows(rows: List[Dict[str, Any]]) -> List[str]:
    """
    Convert [{"video_path": ...}, ...] -> ["pokeagent-speedrun/<video_path>", ...]
    """
    out = []
    for r in rows:
        vp = r.get("video_path", "")
        if vp:
            vp = os.path.join(PATH_PREFIX, vp)
        out.append(vp)
    return out


def _col_letter(n: int) -> str:
    """1-indexed column number to Excel letter"""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def create_retrieval_judging_spreadsheet(
    json_paths: Dict[str, str],
    out_xlsx_path: str = "comparison_output/retrieval_comparison.xlsx",
):
    methods = list(json_paths.keys())
    loaded = {}
    max_len = 0

    # ---- load data ----
    for m in methods:
        rows = _normalize_rows(_load_json_list(json_paths[m]))
        loaded[m] = rows
        max_len = max(max_len, len(rows))

    os.makedirs(os.path.dirname(out_xlsx_path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Judging"

    # ---- headers ----
    headers = ["rank"]
    for m in methods:
        headers.extend([
            f"{m}_video_path",
            f"{m}_good",
            f"{m}_quality_order",
        ])

    ws.append(headers)

    # ---- style header ----
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEECE1")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    # ---- validations ----
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv_bool)

    dv_order = DataValidation(
        type="list",
        formula1=f'"{",".join(str(i) for i in range(1, len(methods) + 1))}"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_order)

    # ---- rows ----
    for i in range(max_len):
        row = [i + 1]
        for m in methods:
            vp = loaded[m][i] if i < len(loaded[m]) else ""
            row.extend([vp, "", ""])
        ws.append(row)

    # ---- apply validations ----
    start_row = 2
    end_row = max_len + 1

    base_col = 2
    block_size = 3

    for mi in range(len(methods)):
        block_start = base_col + mi * block_size
        good_col = block_start + 1
        order_col = block_start + 2

        dv_bool.add(f"{_col_letter(good_col)}{start_row}:{_col_letter(good_col)}{end_row}")
        dv_order.add(f"{_col_letter(order_col)}{start_row}:{_col_letter(order_col)}{end_row}")

    # ---- column widths ----
    ws.column_dimensions["A"].width = 6
    for mi in range(len(methods)):
        block_start = base_col + mi * block_size
        ws.column_dimensions[_col_letter(block_start + 0)].width = 65  # video_path
        ws.column_dimensions[_col_letter(block_start + 1)].width = 10  # good
        ws.column_dimensions[_col_letter(block_start + 2)].width = 18  # quality_order

    wb.save(out_xlsx_path)
    print(f"Wrote spreadsheet: {out_xlsx_path}")


def main():
    # -------------------------
    # Configuration
    # -------------------------
    OUTPUT_DIR = ".cache/pokeagent/xav/comparison_output"
    OUT_XLSX = os.path.join(OUTPUT_DIR, "retrieval_comparison.xlsx")

    JSONS = {
        "max_window_entropy": os.path.join(OUTPUT_DIR, "videos_max_window_entropy.json"),
        "random": os.path.join(OUTPUT_DIR, "videos_random.json"),
        "max_single_match": os.path.join(OUTPUT_DIR, "videos_max_single_match.json"),
        "mean": os.path.join(OUTPUT_DIR, "videos_mean.json"),
    }

    # -------------------------
    # Build spreadsheet
    # -------------------------
    create_retrieval_judging_spreadsheet(
        json_paths=JSONS,
        out_xlsx_path=OUT_XLSX,
    )


if __name__ == "__main__":
    main()
